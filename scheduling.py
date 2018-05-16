# -*- coding: utf-8 -*-
"""
Created on Tue May 15 11:50:37 2018

@author: Zachary Lau
"""

import openpyxl as xl

start_time = 8 #Earliest start time out of 24 hours
divs = 2 #30 minutes per div
end_time = 20 #Latest we can go is 8pm
length = (end_time-start_time)*divs #How many divs we're going to have for a day
days = {0:'Sunday', 1:'Monday', 2:'Tuesday', 3:'Wednesday', 4:'Thursday', 5:'Friday', 6:'Saturday'}

available = []

for i in range(7):
    available.append([])
    for j in range(length):
        available[i].append(0)

def isPM(string):
    if string in ["pm", "PM", "P.M.", "p.m."]:
        return True
    return False

def isAMPM(string):
    if string in ["pm", "PM", "P.M.", "p.m.", "am", "AM", "A.M", "a.m."]:
        return True
    return False
    
def hour_to_index(hour):
#takes in stuff in the formo of 8:15am and spits out an appropriate index, truncating
#Read the hour and minute in
    if len(hour) == 1:
        hourstring = hour
        minutestring = 0
    elif hour[1] == ":":
        hourstring = hour[0]
        minutestring = hour[2:4]
    elif hour[2] == ":":
        hourstring = hour[0:2]
        minutestring = hour[3:5]
    elif hour[1] not in "0123456789":
        hourstring = hour[0]
        minutestring = 0    
    else:
        hourstring = hour[0:2]
        minutestring = 0
    #Check if am or pm
    pm = False
    try:
        if isPM(hour[-2:]):
            pm = True
    except:
        pass
        
    #Now we find the integer value of the index
    
    index = (int(hourstring) - start_time)*divs
    if pm and not hourstring == "12":
        index += 12*divs
    index += int(minutestring)*divs//60
    return index

def index_to_hour(index):
    
    am = True #start by assuming its morning
    hour = start_time + index/divs
    if hour >= 13: #take care of am's and pms
        hour -= 12
        am = False
    minute = int((hour - int(hour))*60)
    
    #now we build the timestring
    
    timestring = str(int(hour)) + ':' + str(minute)
    if timestring[-2] == ":":
        timestring += '0' #take care of single zeroes
    if am:
        timestring += "am"
    else:
        timestring += "pm"
    return timestring

def add_pm(time_pair):
    #time_pair is a tuple of strings. If the first doesn't have a pm and the second does we wil add the pm
    if isPM(time_pair[1][-2:]) and not isAMPM(time_pair[0][-2:]):
        time_pair[0] += "pm"
    return time_pair
        
def read_spreadsheet(filename):
    #takes in the name of a spreadsheet and returns a nested list of availabilities
    times_workbook = xl.load_workbook(filename)
    times_sheet = times_workbook.active #times is the name of the first sheet
    
    current_row = 2 #First row is for labels
    
    while(times_sheet.cell(row = current_row, column = 1).value != None): #Goes through each row starting at 2
        for i in range(7): #goes through each day
            current_column = i+2 #since the columns start at 1 and we don't need the names
            active_cell = times_sheet.cell(row = current_row, column = current_column)
            if active_cell.value != None:
                #We load the values and make them nicer so that they can be read by our index function
                toload = [add_pm(str.strip(time).split("-")) for time in active_cell.value.split(",")]
            else:
                toload = []
            for value in toload:
                start = max(0,hour_to_index(value[0]))
                end = min(hour_to_index(value[1]),length)
                for j in range(start,end):
                    available[i][j] += 1
        current_row += 1
        
    return available

def write_to_spreadsheet(data, filename):
    #data is alist that we would like to write to a spreadsheet filename is the name of the xlsx we write to
    try:
        workbook = xl.load_workbook(filename)
        sheet = workbook.active
        #If we already have that file then we shoul open it
    except:
        workbook = xl.Workbook()
        sheet = workbook.active
        #Otherwise we'l just make a new one
        
    current_row = 1
    current_column = 2
#    This would be if we're adding data to another spreadsheet
#    while(sheet.cell(row = current_row, column = current_column).value != None
#          or sheet.cell(row = current_row + 1, column = current_column +1).value != None):
#        current_row += 1 #so we don't overwrite the data
        
    current_row += 1
    
    start_row = current_row #we're gonna be abusing current_row
    start_column = current_column
    
    for i in range(len(data)):
        try:
            sheet.cell(row = start_row-1,column = start_column+i).value = days[i]
        except:
            pass
        for j in range(len(data[i])):
            
            current_row = start_row+j
            current_column = start_column+i
            sheet.cell(row = current_row, column = current_column).value = data[i][j]
    for j in range(max([len(sub) for sub in data])):
        try:
            sheet.cell(row = start_row + j, column = start_column - 1).value = index_to_hour(j)
        except:
            pass
    
    workbook.save(filename)
    
def main():
    available = read_spreadsheet('times.xlsx')
    write_to_spreadsheet(available, 'timesvisualized.xlsx')

if __name__ == "__main__":
    main()
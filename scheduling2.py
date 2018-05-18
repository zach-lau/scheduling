# -*- coding: utf-8 -*-
"""
Created on Wed May 16 16:34:19 2018

@author: Zachary Lau
"""

import openpyxl as xl
import functions

#--------------------------------------------

def read(filename):
    #filename: name of file containing data
    #returns: any data in the file in a 2d list. Will start going down the left hand side until it reaches a blank space
    #Assumes the first column and row are used for labels
    
    #Define the variables that will be used throughout
    
    data = []
    workbook = xl.load_workbook(filename).active
    
    #Helper functions --------------------------------------------
    
    def get_width(workbook):
        #returns the width of the data in the file
        n = 0
        while workbook.cell(row = 1, column = n+1).value != None:
            n += 1
        return n
    
    def get_height(workbook):
        #returns the height of the data in the file
        n = 0
        while workbook.cell(row = n+1, column = 1).value != None:
            n += 1
        return n
    
    #Actual work-----------------------------------------------------
    
    width = get_width(workbook)
    height = get_height(workbook)
    
    for i in range(width):
        data.append([])
        for j in range(height):
            data[i].append(workbook.cell(row = i + 2, column = j + 2).value)
    
    return data
        

def analyze(data, function):
    #data: data to be analyzed
    #returns: 2d list that is the output of the analysis function
    return function(data)
    

def write(filename, data, start_row, start_column):
    #filename: name of file to be written to
    #data: 2d list of data to be written
    #start_row: row number of top left corner
    #start_column: column number of top left corner
    #returns: None
    workbook = xl.load_workbook(filename).active
    
    if data == None:
        return
    for d_row in len(data):
        for d_column in len(data[d_row]):
            workbook.cell(row = start_row + d_row, column = start_column + d_column).value = data[d_row][d_column]
    workbook.save()
    
    
#-----------------------------------------

def main(fileIn, fileOut, function):
    #fileIn: file that data will be read from
    #fileOut: file that data will be written to
    #return: None. Writes analyzed data to fileOut
    dataIn = read(fileIn)
    for row in dataIn:
        print(row)
    dataOut = analyze(dataIn, function)
    write(fileOut, dataOut, 1, 1)
    
if __name__ == "__main__":
    main("times.xlsx", "times1.xlsx", functions.times)